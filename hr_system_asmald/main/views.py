from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse, HttpResponseNotAllowed, HttpResponseNotFound
from django.shortcuts import render, redirect
from django.views.generic import DetailView

from .form import StaffForm, TodoForm
from .models import Anketa, Users, Work, Staff, Department, Position, ToDo, ExecutionTime
from .eskiz import SendSmsApiWithEskiz
from django.shortcuts import get_object_or_404
from django.utils import timezone
from datetime import date
from datetime import datetime, time
import xlwt
import datetime
import pytz
from django.http import HttpResponse
from openpyxl import Workbook
from django.core.paginator import Paginator
from django.db.models import Q
from django.views.decorators.csrf import csrf_exempt


@login_required(login_url='/login/')
def data(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="staff.xls"'.format(
            timezone.datetime.now().strftime('%Y%m%d'))

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Ishchilar royxati')

    row_num = 0

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    columns = ['Shartnoma raqami','F.I.SH','Tugilgan sana','Tel raqami','Manzil','Department', 'Position', 'Jinsi', 'Kegan Vaqti', 'Comment','Status']

    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_style)

    # Sheet body, remaining rows
    font_style = xlwt.XFStyle()
    font_style.num_format_str = 'dd/mm/yyyy'

    rows = Staff.objects.all().values_list('contract_id', 'full_name', 'b_date', 'phone', 'address', 'department__department_name', 'position__position_name', 'gender', 'date_join', 'comment', 'status')

    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num], font_style)

    wb.save(response)
    return response


@login_required(login_url='/login/')
def activ_data(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="active_staff.xls"'.format(
        timezone.datetime.now().strftime('%Y%m%d'))

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Ishchilar royxati')

    row_num = 0

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    columns = ['Shartnoma raqami','F.I.SH', 'Tugilgan sana', 'Tel raqami', 'Manzil', 'Department', 'Position', 'Jinsi', 'Kegan Vaqti', 'Comment',
               'Status']

    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_style)

    # Sheet body, remaining rows
    font_style = xlwt.XFStyle()
    font_style.num_format_str = 'dd/mm/yyyy'

    rows = Staff.objects.filter(status=True).values_list('contract_id', 'full_name', 'b_date', 'phone', 'address', 'department__department_name',
                                           'position__position_name', 'gender', 'date_join', 'comment', 'status')

    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num], font_style)

    wb.save(response)
    return response



@login_required(login_url='/login/')
def false_data(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="dismissed_staff.xls"'.format(
        timezone.datetime.now().strftime('%Y%m%d'))

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Ishchilar royxati')

    row_num = 0

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    columns = ['Shartnoma raqami','F.I.SH', 'Tugilgan sana', 'Tel raqami', 'Manzil', 'Department', 'Position', 'Jinsi', 'Kegan Vaqti', 'Comment',
               'Status']

    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_style)

    # Sheet body, remaining rows
    font_style = xlwt.XFStyle()
    font_style.num_format_str = 'dd/mm/yyyy'

    rows = Staff.objects.filter(status=False).values_list('contract_id','full_name', 'b_date', 'phone', 'address', 'department__department_name',
                                           'position__position_name', 'gender', 'date_join', 'comment', 'status')
    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num], font_style)

    wb.save(response)
    return response


@login_required(login_url='/login/')
def anketa_data(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="anketa.xlsx"'.format(
        datetime.datetime.now().strftime('%Y%m%d'))

    wb = Workbook()
    ws = wb.active
    ws.title = 'Bot orqali qabul qilinganlar'

    row_num = 1

    columns = ['F.I.SH','Tugilgan sana','Tel raqami','Hudud','Manzil','Talim darajasi','Eski ish joyi','Position', 'Qoshimcha', 'Vaqti']

    for col_num, column_title in enumerate(columns, 1):
        ws.cell(row=row_num, column=col_num, value=column_title)

    rows = Anketa.objects.all().values_list('full_name', 'b_date', 'phone', 'region', 'address', 'education', 'old_work', 'position', 'additions', 'create_at')

    for row_data in rows:
        row_num += 1
        for col_num, value in enumerate(row_data, 1):
            if isinstance(value, datetime.datetime):
                value = value.replace(tzinfo=pytz.UTC).astimezone(pytz.timezone('UTC')).replace(tzinfo=None)
            ws.cell(row=row_num, column=col_num, value=value)

    wb.save(response)
    return response



@login_required(login_url='/login/')
def rezerv_data(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="rezerv.xlsx"'.format(
        datetime.datetime.now().strftime('%Y%m%d'))

    wb = Workbook()
    ws = wb.active
    ws.title = 'Bot orqali qabul qilinganlar'

    row_num = 1

    columns = ['F.I.SH','Tugilgan sana','Tel raqami','Hudud','Manzil','Talim darajasi','Eski ish joyi','Position', 'Qoshimcha', 'Vaqti']

    for col_num, column_title in enumerate(columns, 1):
        ws.cell(row=row_num, column=col_num, value=column_title)

    rows = Anketa.objects.filter(rezerv=True).values_list('full_name', 'b_date', 'phone', 'region', 'address', 'education', 'old_work', 'position', 'additions', 'create_at')

    for row_data in rows:
        row_num += 1
        for col_num, value in enumerate(row_data, 1):
            if isinstance(value, datetime.datetime):
                value = value.replace(tzinfo=pytz.UTC).astimezone(pytz.timezone('UTC')).replace(tzinfo=None)
            ws.cell(row=row_num, column=col_num, value=value)

    wb.save(response)
    return response


#
# def get_birthdays_this_week():
#     today = date.today()
#     monday = today - datetime.timedelta(days=today.weekday())  # Get the Monday of the current week
#     sunday = monday + datetime.timedelta(days=6)  # Get the Sunday of the current week
#     birthdays = Staff.objects.filter(b_date__month=today.month, b_date__range=[monday, sunday])
#     return birthdays


@login_required(login_url='/login/')
def index(request):
    count_user = Users.objects.count()
    count_work = Work.objects.count()
    count_anketa = Anketa.objects.count()
    count_rezerv = Anketa.get_rezerv_objects().count()
    count_false = Anketa.get_false().count()
    staff_true = Staff.get_true().count()
    staff_false = Staff.get_false().count()
    count_staff = Staff.objects.all().count()
    count_suhbat = Anketa.get_selected_objects().count()
    today = date.today()
    count_bday = Staff.objects.filter(status=True, b_date__day=today.day, b_date__month=today.month).count()
    staff = Staff.objects.all()
    todo_items = ToDo.objects.order_by('id')[::-1]
    count_weekly_b = Staff.get_birthdays_between_monday_and_sunday().count()
    paginator = Paginator(todo_items, 5)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    count_male = Staff.objects.filter(gender='Erkak').count()
    count_famale = Staff.objects.filter(gender='Ayol').count()
    chart_list = ['Ketganlar', 'Ishlab turganlar', 'Rezervdagilar']
    chart_number = [staff_false, staff_true, count_rezerv]

    start_date_30 = today.replace(year=today.year - 50)
    end_date_30 = today.replace(year=today.year - 31)
    start_date_18 = today.replace(year=today.year - 30)
    end_date_18 = today.replace(year=today.year - 18)
    start_date_80 = today.replace(year=today.year-80)
    end_date_80 = today.replace(year=today.year-51)
    staff_between_30_and_50 = Staff.objects.filter(b_date__range=[start_date_30, end_date_30])
    staff_between_18_and_30 = Staff.objects.filter(b_date__range=[start_date_18, end_date_18])
    staff_between_51_and_80 = Staff.objects.filter(b_date__range=[start_date_80, end_date_80])
    total_staff_count_30_to_50 = staff_between_30_and_50.count()
    total_staff_count_18_to_30 = staff_between_18_and_30.count()
    total_staff_count_51_to_80 = staff_between_51_and_80.count()

    gender_list = ['Erkak', 'Ayol', '18-30', '30-50', '50 dan tepa']
    gender_number = [count_male, count_famale, total_staff_count_18_to_30, total_staff_count_30_to_50, total_staff_count_51_to_80]
    context = {
        'count_user': count_user,
        'count_work': count_work,
        'count_anketa': count_anketa,
        'count_rezerv': count_rezerv,
        'count_false': count_false,
        'count_staff': count_staff,
        'count_suhbat': count_suhbat,
        'staff_true': staff_true,
        'staff_false': staff_false,
        'count_bday': count_bday,
        'staff': staff,
        'todo_items': todo_items,
        'count_weekly_b': count_weekly_b,
        'page_obj': page_obj,
        'gender_list': gender_list,
        'gender_number': gender_number,
        'chart_list': chart_list,
        'chart_number': chart_number
    }
    return render(request, 'index.html', context)



@login_required(login_url='/login/')
def anketa(request):
    if request.method == 'POST':
        select_value = request.POST.get('select')
        anketa_id = request.POST.get('anketa_id')
        try:
            anketa = Anketa.objects.get(id=anketa_id)
            if select_value == 'true':
                anketa.select = True
            else:
                anketa.select = False
            anketa.save()
            return redirect('anketa')
        except Anketa.DoesNotExist:
            return HttpResponse('Anketa not found.')
    else:
        anketa_list = Anketa.objects.all()

    if 'q' in request.GET:
        q = request.GET['q']
        anketa_list = Anketa.objects.filter(Q(full_name__icontains=q) | Q(phone__icontains=q))
    else:
        anketa_list = Anketa.objects.all()
    paginator = Paginator(anketa_list, 5)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {
        'anketa_list': anketa_list,
        'page_obj': page_obj
    }
    return render(request, 'anketa_all.html', context)



@login_required(login_url='/login/')
def select_anketa(request):
    if request.method == 'POST':
        select_value = request.POST.get('select')
        select_id = request.POST.get('select_id')
        try:
            anketa = Anketa.objects.get(id=select_id)
            print(anketa)
            name = anketa.full_name
            phone = Anketa.objects.filter(id=select_id).values('phone')
            if select_value == 'sms':
                text_sms = f'Salom hurmatli {name}. Sizni qoldirgan anketangiz sabab sizni suhbatga taklif qilamiz.'
                sms = SendSmsApiWithEskiz(text_sms, int(phone[0].get('phone')))
                sms.send()
            elif select_value == 'true':
                anketa.rezerv = True
            else:
                anketa.rezerv = False
            anketa.save()
            return redirect('select')
        except Anketa.DoesNotExist:
            return HttpResponse('Anketa not found.')
    else:

        selected_objects = Anketa.get_selected_objects()
        paginator = Paginator(selected_objects, 5)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        context = {
            'selected_objects': selected_objects,
            'page_obj': page_obj
        }
        return render(request, 'selected.html', context)



@login_required(login_url='/login/')
def staff(request):
    if 'q' in request.GET:
        q = request.GET['q']
        staffs = Staff.objects.filter(Q(full_name__icontains=q) | Q(phone__icontains=q))
    else:
        staffs = Staff.objects.all()
    paginator = Paginator(staffs, 5)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'staffs': staffs,
        'page_obj': page_obj
    }
    return render(request, 'worker.html', context)

@login_required(login_url='/login/')
def detail_staff(request):
    return render(request, 'worker_detail.html')


class StaffDetailView(DetailView):
    model = Staff
    template_name = 'worker_detail.html'


@login_required(login_url='/login/')
def active_staff(request):
    if 'q' in request.GET:
        q = request.GET['q']
        active = Staff.objects.filter(Q(full_name__icontains=q) | Q(phone__icontains=q), status=True)
    else:
        active = Staff.get_true()
    paginator = Paginator(active, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {
        'active': active,
        'page_obj': page_obj,
    }
    return render(request, 'active_staff.html', context)


@login_required(login_url='/login/')
def false_staff(request):
    if 'q' in request.GET:
        q = request.GET['q']
        false = Staff.objects.filter(Q(full_name__icontains=q) | Q(phone__icontains=q), status=False)
    else:
        false = Staff.get_false()
    paginator = Paginator(false, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {
        'false': false,
        'page_obj': page_obj
    }
    return render(request, 'false_staff.html', context)


@login_required(login_url='/login/')
def send_sms(request, staff_id):
    staff = Staff.objects.get(id=staff_id)
    phone = Staff.objects.filter(id=staff_id).values('phone')
    if request.method == 'POST':
        message = request.POST.get('habar')
        name = staff.full_name
        text = f"{name} {message}\nXurmat bilan EUROPRINT jamoasi"
        sms = SendSmsApiWithEskiz(text, int(phone[0].get('phone')))
        sms.send()
        return redirect('staffs')
    else:
        return render(request, 'send_sms.html', {'staff': staff})



@login_required(login_url='/login/')
def send_message(request):
    if request.method == 'POST':
        message = request.POST.get('habar')
        staffs = Staff.objects.all()
        for staff in staffs:
            phone = int(staff.phone)
            text = f"Assalom alekum xurmatli jamoamiz a'zosi {staff.full_name}. {message}. Xurmat bilan EUROPRINT jamoasi."
            sms = SendSmsApiWithEskiz(text, phone)
            sms.send()
        return redirect('index')
    else:
        return render(request, 'send_message_all.html')



# def today_birthday(request):
#     today = timezone.now().date()
#     staffs = Staff.objects.filter(status=True, b_date__day=today.day, b_date__month=today.month)
#     for staff in staffs:
#         text = f'Assalomu alaykum {staff.full_name}. Sizni tavallud ayyomingiz bilan tabriklaymiz.\n Xurmat bilan EUROPRINT'
#         phone = staff.phone
#         print(phone)
#         sms = SendSmsApiWithEskiz(text, int(phone))
#         sms.send()
#     return render(request, 'today_bithday.html', {'staffs': staffs})



@login_required(login_url='/login/')
def today_birthday(request):
    today = timezone.now().date()
    staffs = Staff.objects.filter(status=True, b_date__day=today.day, b_date__month=today.month)
    sent_phones = set()
    for staff in staffs:
        text = f'Assalomu alaykum {staff.full_name}. Sizni tavallud ayyomingiz bilan tabriklaymiz.\n Xurmat bilan EUROPRINT'
        phone = staff.phone
        if phone not in sent_phones:
            sms = SendSmsApiWithEskiz(text, int(phone))
            sms.send()
            sent_phones.add(phone)
    return render(request, 'today_bithday.html', {'staffs': staffs})



@login_required(login_url='/login/')
def weekly_birthday(request):
    weekly_birthday = Staff.get_birthdays_between_monday_and_sunday()
    print(weekly_birthday)
    context = {
        'weekly_birthday': weekly_birthday
    }
    return render(request, 'weekly_birthday.html', context)



@login_required(login_url='/login/')
def delete_staff(request, staff_id):
    if request.method == 'POST':
        staff = get_object_or_404(Staff, id=staff_id)
        staff.delete()
        return redirect('staffs')
    else:
        return HttpResponseNotAllowed(['POST'])


@login_required(login_url='/login/')
def edit_staff(request, staff_id):
    staff = get_object_or_404(Staff, id=staff_id)
    name = staff.full_name
    departments = Department.objects.all()
    positions = Position.objects.all()

    if request.method == 'POST':
        form = StaffForm(request.POST, instance=staff)
        if form.is_valid():
            form.save()
            return redirect('staffs')
    else:
        form = StaffForm(instance=staff)
    context = {
        'departments': departments,
        'form': form,
        'positions': positions,
        'name': name,
        'staff': staff

    }
    return render(request, 'edit_staff.html', context)



@login_required(login_url='/login/')
def add_staff(request):
    departments = Department.objects.all()
    positions = Position.objects.all()

    if request.method == 'POST':
        form = StaffForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('index')
    else:
        form = StaffForm()
    context = {
        'departments': departments,
        'form': form,
        'positions': positions,
        'staff': staff,

    }
    return render(request, 'add_staff.html', context)



@login_required(login_url='/login/')
def rezerv(request):
    if request.method == 'POST':
        rezerv_value = request.POST.get('rezerv')
        rezerv_id = request.POST.get('rezerv_id')
        try:
            anketa = Anketa.objects.get(id=rezerv_id)
            name = anketa.full_name
            phone = Anketa.objects.filter(id=rezerv_id).values('phone')
            if rezerv_value == 'false':
                anketa.rezerv = False
                anketa.save()
            elif rezerv_value == 'sms':
                text_sms = f'Salom hurmatli {name}. Sizni qoldirgan anketangiz sabab sizni suhbatga taklif qilamiz.'
                sms = SendSmsApiWithEskiz(text_sms, int(phone[0].get('phone')))
                sms.send()
            anketa.save()
            return redirect('rezervs')
        except Anketa.DoesNotExist:
            return HttpResponse('Rezerv not found.')
    else:
        rezervs = Anketa.get_rezerv_objects()
        paginator = Paginator(rezervs, 10)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        context = {
            'rezervs': rezervs,
            'page_obj': page_obj
        }

        return render(request, 'rezerv_all.html', context)


@login_required(login_url='/login/')
def edit_todo(request, todo_id):
    task = ToDo.objects.all()
    todo = ToDo.objects.get(id=todo_id)
    if request.method == 'POST':
        form = TodoForm(request.POST, instance=todo)
        if form.is_valid():
            form.save()
            return redirect('index')
    else:
        form = TodoForm(instance=todo)

    return render(request, 'edit_todo.html', {'form': form, 'todo': todo, 'task': task})



@login_required(login_url='/login/')
def add_todo(request):
    form = TodoForm()
    if request.method == 'POST':
        form = TodoForm(request.POST)
        if form.is_valid():
            task = form.cleaned_data['task']
            ToDo.objects.create(task=task)
            return redirect('index')

    return render(request, 'add_todo.html', {'form': form})


@login_required(login_url='/login/')
def delete_todo(request, item_id):
    task = ToDo.objects.get(id=item_id)
    task.delete()
    return redirect('index')


@login_required(login_url='/login/')
def work(request):
    work_all = Work.objects.all()
    context = {

        'work_all': work_all,
    }
    return render(request, 'work.html', context)


@login_required(login_url='/login/')
def delete_work(request, item_id):
    work = Work.objects.get(id=item_id)
    work.delete()
    return redirect('work')


@login_required(login_url='/login/')
def user(request):
    user_all = Users.objects.all()
    context = {
        'user_all': user_all
    }
    return render(request, 'users.html', context)



@login_required(login_url='/login/')
def delete_user(request, item_id):
    user = Users.objects.get(id=item_id)
    user.delete()
    return redirect('index')


