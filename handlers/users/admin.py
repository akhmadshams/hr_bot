import asyncio
from aiogram import types, utils
from aiogram.dispatcher import FSMContext
from aiogram.types import Message, ReplyKeyboardRemove
from data.config import ADMINS
from loader import dp, db, bot
from keyboards.default.admin_key import admin_menu
from keyboards.default.menu import main_menu
import sqlite3
from openpyxl import Workbook
from io import BytesIO
from datetime import datetime, timedelta
from datetime import date


connection = sqlite3.connect('data/main.db')
cursor = connection.cursor()


@dp.message_handler(commands='admin', user_id=ADMINS)
async def admins(message:Message):
    await message.answer('<b>admin_handler menu ochildi qaytish uchun 🎯 Bosh menu tugmasini bosing</b>',reply_markup=admin_menu)


@dp.message_handler(text='🎯 Bosh menu', user_id=ADMINS)
async def main_men(message: Message):
    await message.answer("<b>Bosh menu</b>", reply_markup=main_menu)




@dp.message_handler(text='📈 Umumiy ma\'lumot')
async def all_date(message: Message):
    #malumotlarni hammasini olish
    cursor.execute("SELECT * FROM Anketa")
    rows = cursor.fetchall()

    #excel faylga endi yozamiz malumotlarni
    wb = Workbook()
    ws = wb.active
    ws.title ='Umumiy malumotlar'

    ws.cell(row=1, column=1).value='ID'
    ws.cell(row=1, column=2).value='Ism Fam'
    ws.cell(row=1, column=3).value='Tugilgan sana'
    ws.cell(row=1, column=4).value='Tel raqami'
    ws.cell(row=1, column=5).value='Hudud'
    ws.cell(row=1, column=6).value='Manzil'
    ws.cell(row=1, column=7).value='Malumoti'
    ws.cell(row=1, column=8).value='Eski ish joyi'
    ws.cell(row=1, column=9).value='Positsiyasi'
    ws.cell(row=1, column=10).value='Qoshimcha'
    ws.cell(row=1, column=11).value='Royxatdan otgan sana'

    for row in rows:
        ws.append(row)

    #file ni yuborish
    file = BytesIO()
    wb.save(file)
    file.seek(0)
    document = types.InputFile(file, filename='umumiy.xlsx')
    await message.answer_document(document=document)
    wb.close()


@dp.message_handler(text='📈 Oylik ma\'lumotlar')
async def monthly(message: Message):
    wb = Workbook()
    ws = wb.active
    sana = datetime.now().date() - timedelta(days=datetime.now().day - 1)
    query = f"SELECT * FROM Anketa WHERE created_at >= '{sana}'"
    cursor.execute(query)
    data = cursor.fetchall()

    #faylga malumotni yozish
    ws.title = 'Oylik malumot'

    ws.cell(row=1, column=1).value='ID'
    ws.cell(row=1, column=2).value='Ism Fam'
    ws.cell(row=1, column=3).value='Tugilgan sana'
    ws.cell(row=1, column=4).value='Tel raqami'
    ws.cell(row=1, column=5).value='Hudud'
    ws.cell(row=1, column=6).value='Manzil'
    ws.cell(row=1, column=7).value='Malumoti'
    ws.cell(row=1, column=8).value='Eski ish joyi'
    ws.cell(row=1, column=9).value='Positsiyasi'
    ws.cell(row=1, column=10).value='Qoshimcha'
    ws.cell(row=1, column=11).value='Royxatdan otgan sana'

    for row in data:
        ws.append(row)

    filename = f'malumotlar_{sana:%Y-%m}.xlsx' #bu bizga fayl yuborilganda fayl nomida yil va oy berilgan boladi
    wb.save(filename)
    file = BytesIO()
    wb.save(file)
    file.seek(0)
    document = types.InputFile(file, filename=filename)
    await message.answer_document(document=document)
    wb.close()

    # boldi oylik malumotni ham bitirib oldik


@dp.message_handler(text='📈 Haftalik ma\'lumotlar')
async def weekly(message: Message):
    wb = Workbook()
    ws = wb.active
    query = f"SELECT * FROM Anketa WHERE created_at >= date('now', '-7 day')"
    cursor.execute(query)
    data = cursor.fetchall()

    #faylga malumotni yozish
    ws.title = 'Haftalik malumotlar'

    ws.cell(row=1, column=1).value='ID'
    ws.cell(row=1, column=2).value='Ism Fam'
    ws.cell(row=1, column=3).value='Tugilgan sana'
    ws.cell(row=1, column=4).value='Tel raqami'
    ws.cell(row=1, column=5).value='Hudud'
    ws.cell(row=1, column=6).value='Manzil'
    ws.cell(row=1, column=7).value='Malumoti'
    ws.cell(row=1, column=8).value='Eski ish joyi'
    ws.cell(row=1, column=9).value='Positsiyasi'
    ws.cell(row=1, column=10).value='Qoshimcha'
    ws.cell(row=1, column=11).value='Royxatdan otgan sana'

    for row in data:
        ws.append(row)

    filename = 'haftalik.xlsx' #bu bizga fayl yuborilganda fayl nomida yil va oy berilgan boladi
    wb.save(filename)
    file = BytesIO()
    wb.save(file)
    file.seek(0)
    document = types.InputFile(file, filename=filename)
    await message.answer_document(document=document)
    wb.close()



@dp.message_handler(text='📈 Kunlik ma\'lumotlar')
async def weekly(message: Message):
    wb = Workbook()
    ws = wb.active
    query = f"SELECT * FROM Anketa WHERE created_at >= date('now', '-0 day')"
    cursor.execute(query)
    data = cursor.fetchall()

    #faylga malumotni yozish
    ws.title = 'Haftalik malumotlar'

    ws.cell(row=1, column=1).value='ID'
    ws.cell(row=1, column=2).value='Ism Fam'
    ws.cell(row=1, column=3).value='Tugilgan sana'
    ws.cell(row=1, column=4).value='Tel raqami'
    ws.cell(row=1, column=5).value='Hudud'
    ws.cell(row=1, column=6).value='Manzil'
    ws.cell(row=1, column=7).value='Malumoti'
    ws.cell(row=1, column=8).value='Eski ish joyi'
    ws.cell(row=1, column=9).value='Positsiyasi'
    ws.cell(row=1, column=10).value='Qoshimcha'
    ws.cell(row=1, column=11).value='Royxatdan otgan sana'
    for row in data:
        ws.append(row)

    filename = 'kunlik.xlsx' #bu bizga fayl yuborilganda fayl nomida yil va oy berilgan boladi
    wb.save(filename)
    file = BytesIO()
    wb.save(file)
    file.seek(0)
    document = types.InputFile(file, filename=filename)
    await message.answer_document(document=document)
    wb.close()



@dp.message_handler(text="🎯 Reklama", user_id=ADMINS, state=None)
async def send_ad_to_all(message: types.Message, state=FSMContext):
    try:
        await message.answer(
            "✅ Reklamani botga forward qiling. Siz forward qilgan reklama to'gridan to'g'ri barcha foydalanuchilarga yuboriladi. \n\n✅ Yoki xabar matnini kiriting: \n\n⚠️ Xabar yuborishni istamasangiz /bekor kamandasini kiriting.",
            reply_markup=ReplyKeyboardRemove())
        # print("1")
    except Exception as e:
        await message.answer(
            "✅ Reklamani botga forward qiling. Siz forward qilgan reklama to'gridan to'g'ri barcha foydalanuchilarga yuboriladi. \n\n✅ Yoki xabar matnini kiriting: \n\n⚠️ Xabar yuborishni istamasangiz /bekor kamandasini kiriting.")
        # print("2")

    await state.set_state("send_users")


@dp.message_handler(state="send_users", text="/bekor")
async def send_ad_to_all(message: types.Message, state=FSMContext):
    await message.answer("❌ Bekor qilindi", reply_markup=admin_menu)
    await state.finish()


@dp.message_handler(state="send_users", content_types="text", is_forwarded=False)
async def send_ad_to_all(message: types.Message, state=FSMContext):
    await state.finish()
    bloklaganlar = 0
    jonlilar = 0

    users = db.select_all_users()
    for user in users:
        user_id = user[0]
        try:
            await bot.send_message(user_id, message.text)
            jonlilar += 1

        except utils.exceptions.BotBlocked as e:
            bloklaganlar += 1

        except:
            pass

        else:
            pass
        await asyncio.sleep(0.05)
    await message.answer("✅ Xabaringiz foydalanuvchilarga yetkazildi...", reply_markup=admin_menu)
    await message.answer(f"Jami foydalanuvchilar soni: {bloklaganlar + jonlilar}")
    await message.answer(f"Bloklagan foydalanuvchilar soni: {bloklaganlar}")
    await message.answer(f"Faol foydalanuvchilar soni: {jonlilar}")


@dp.message_handler(state="send_users", content_types="any", is_forwarded=True)
async def send_ad_to_all(message: types.Message, state=FSMContext):
    await state.finish()

    bloklaganlar = 0
    jonlilar = 0

    users = db.select_all_users()
    for user in users:
        user_id = user[0]

        try:
            await message.forward(user_id, message.forward_from.id, message.forward_from_message_id)
            jonlilar += 1
        except utils.exceptions.BotBlocked as e:
            bloklaganlar += 1

        except:
            await message.forward(user_id, message.forward_from_chat.id, message.forward_from_message_id)
            jonlilar += 1
        else:
            pass

        await asyncio.sleep(0.05)

    await message.answer("✅ Xabaringiz foydalanuvchilarga yetkazildi...")

    await message.answer(f"Jami foydalanuvchilar soni: {bloklaganlar + jonlilar}")
    await message.answer(f"Bloklagan foydalanuvchilar soni: {bloklaganlar}")
    await message.answer(f"Faol foydalanuvchilar soni: {jonlilar}")


@dp.message_handler(text="🔄 Foydalanuvchilar soni", user_id=ADMINS)
async def get_all_users(message: types.Message):
    today = date.today()
    current_time = today.strftime("%B %d, %Y")
    users = db.count_users()
    text = f"Jami foydalanuvchilar soni<b> {users[0]} </b> ta  - <b>{current_time}</b> holatiga"
    await message.answer(text)


@dp.message_handler(text="/cleandb", user_id=ADMINS)
async def get_all_users(message: types.Message):
    db.delete_users()
    await message.answer("Baza tozalandi!")